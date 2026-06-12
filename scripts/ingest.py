#!/usr/bin/env python3
"""
WC26 Quiniela — ingestion from the single organiser workbook.

Reads /Users/edsherrington/Downloads/WC26 Quiniela ORGANISER.xlsx and produces:
  ../data/fixtures.json     (group-stage fixture list + knockout round metadata)
  ../data/predictions.json  (roster + every entrant's picks)

Sheets used:
  'Actual Results'    — fixture dates, venues, teams (Groups A–L, 6 matches each)
  'Raw Data & Calcs'  — all entrant picks, one row per pick, flat table

Run:  python3 scripts/ingest.py   (from the project root)
"""
import json
import re
import os
import unicodedata
# Kick-off times are hardcoded BST in data/kickoffs.json (keyed by sorted team-code
# pair, e.g. "ARG|JOR"), sourced from the Sky Sports schedule. We no longer derive
# them from venue timezones — that local->UTC->BST round-trip produced wrong times.

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA_DIR = os.path.join(ROOT, "data")

# Hardcoded BST kick-offs, keyed by sorted team-code pair (e.g. "ARG|JOR").
with open(os.path.join(DATA_DIR, "kickoffs.json")) as _kf:
    KICKOFFS = json.load(_kf)["kickoffs"]

ORGANISER_PATH = os.path.expanduser(
    "~/Downloads/WC26 Quiniela ORGANISER.xlsx"
)

KO_ROUNDS = [
    {"id": "R32",    "label": "Round of 32",   "points": 2,  "prefix": "R32-"},
    {"id": "R16",    "label": "Round of 16",   "points": 4,  "prefix": "R16-"},
    {"id": "QF",     "label": "Quarter-final", "points": 6,  "prefix": "QF-"},
    {"id": "SF",     "label": "Semi-final",    "points": 9,  "prefix": "SF-"},
    {"id": "F",      "label": "Finalist",      "points": 14, "prefix": "F-"},
    {"id": "WINNER", "label": "Champion",      "points": 25, "prefix": "WINNER"},
]

# Map spreadsheet stage labels to round IDs
STAGE_TO_ROUND = {
    "round of 32":  "R32",
    "round of 16":  "R16",
    "quarter-final":"QF",
    "semi-final":   "SF",
    "finalist":     "F",
    "champion":     "WINNER",
}

GROUP_POINTS = {"exact": 4, "result_gd": 2, "result": 1, "wrong": 0}


def slugify(name):
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s


def parse_team(s):
    """'Mexico (MEX)' -> ('Mexico', 'MEX'); 'USA' -> ('USA', 'USA')."""
    if s is None:
        return None, None
    s = str(s).strip()
    m = re.match(r"^(.*?)\s*\(([A-Z]{3})\)\s*$", s)
    if m:
        return m.group(1).strip(), m.group(2)
    return s, s


def as_int(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def build_fixtures(wb):
    """Build 72 group-stage fixtures from the 'Actual Results' sheet."""
    ws = wb["Actual Results"]
    fixtures = []
    current_group = None
    n = 0
    for row in ws.iter_rows(values_only=True):
        b, c, d, _, f = (row[1], row[2], row[3], row[4], row[5])
        # Group header row: col B is a string like "Group A"
        if isinstance(b, str) and b.strip().lower().startswith("group "):
            current_group = b.strip().split()[-1]
            continue
        # Fixture row: col B is a datetime, col D and F are team strings
        if d and f and not isinstance(b, str) and b is not None:
            n += 1
            home_name, home_code = parse_team(d)
            away_name, away_code = parse_team(f)
            pair = "|".join(sorted((home_code, away_code)))
            kickoff = KICKOFFS.get(pair)
            if kickoff is None:
                print(f"  WARN: no hardcoded BST kick-off for {home_code} v {away_code}")
            fixtures.append({
                "id": f"GS-{n:02d}",
                "group": current_group,
                "kickoff": kickoff,
                "venue": str(c).strip() if c else None,
                "home": {"name": home_name, "code": home_code},
                "away": {"name": away_name, "code": away_code},
            })
    return fixtures


def build_entrants(wb):
    """Build all entrants from the 'Raw Data & Calcs' sheet."""
    ws = wb["Raw Data & Calcs"]
    entrants_map = {}  # name -> entrant dict (in insertion order)

    for row in ws.iter_rows(min_row=2, values_only=True):
        entrant_name = row[0]
        if not entrant_name or not str(entrant_name).strip():
            continue
        name = str(entrant_name).strip()
        team = str(row[1]).strip() if row[1] else ""
        stage = str(row[2]).strip() if row[2] else ""
        pick_id = str(row[4]).strip() if row[4] else ""
        a_goals = as_int(row[6])
        b_goals = as_int(row[8])
        chosen = str(row[9]).strip() if row[9] else None

        if name not in entrants_map:
            slug = slugify(name)
            logo_rel = f"assets/logos/{slug}.png"
            logo_abs = os.path.join(ROOT, logo_rel)
            entrants_map[name] = {
                "id": slug,
                "name": name,
                "teamName": team,
                "logo": logo_rel if os.path.exists(logo_abs) else None,
                "groupStage": {},
                "knockout": {"R32": [], "R16": [], "QF": [], "SF": [], "F": [], "WINNER": None},
            }

        e = entrants_map[name]

        if stage.lower() == "group stage":
            if pick_id.startswith("GS-"):
                e["groupStage"][pick_id] = [a_goals, b_goals]
        else:
            round_id = STAGE_TO_ROUND.get(stage.lower())
            if round_id and chosen:
                if round_id == "WINNER":
                    e["knockout"]["WINNER"] = chosen
                else:
                    e["knockout"][round_id].append(chosen)

    return list(entrants_map.values())


def main():
    if not os.path.exists(ORGANISER_PATH):
        raise SystemExit(f"Organiser file not found: {ORGANISER_PATH}")

    import openpyxl
    print(f"Reading: {ORGANISER_PATH}")
    wb = openpyxl.load_workbook(ORGANISER_PATH, data_only=True)

    fixtures = build_fixtures(wb)
    print(f"  {len(fixtures)} group-stage fixtures")

    entrants = build_entrants(wb)

    for e in entrants:
        gs_count = len(e["groupStage"])
        miss = "" if e["logo"] else "  (no logo yet)"
        print(f"  + {e['name']} — \"{e['teamName']}\" ({gs_count} GS picks){miss}")

    fixtures_doc = {
        "competition": "FIFA World Cup 2026",
        "groupPoints": GROUP_POINTS,
        "knockoutRounds": KO_ROUNDS,
        "groupStage": fixtures,
    }
    predictions_doc = {"entrants": entrants}

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(os.path.join(DATA_DIR, "fixtures.json"), "w") as fh:
        json.dump(fixtures_doc, fh, indent=2, ensure_ascii=False)
    with open(os.path.join(DATA_DIR, "predictions.json"), "w") as fh:
        json.dump(predictions_doc, fh, indent=2, ensure_ascii=False)

    print(f"\nWrote {len(fixtures)} fixtures and {len(entrants)} entrants.")
    print("  data/fixtures.json")
    print("  data/predictions.json")


if __name__ == "__main__":
    main()
